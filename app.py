import os
import json
import requests
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from flask import Flask, request, jsonify, send_file, send_from_directory
import gspread
import logging
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import tempfile

app = Flask(__name__)

# 加载 .env 文件
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Google Sheets API configuration
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/drive.file']
SPREADSHEET_ID = os.getenv('GOOGLE_SHEETS_ID')  # 从环境变量中读取 Google Sheet ID

# 获取环境变量
GOOGLE_APPLICATION_CREDENTIALS = 'credentials/service_account_credentials.json'

# 检查文件是否存在
if not os.path.exists(GOOGLE_APPLICATION_CREDENTIALS):
    raise FileNotFoundError(f"File {GOOGLE_APPLICATION_CREDENTIALS} not found")

# 从文件加载 JSON 格式的服务账户密钥
with open(GOOGLE_APPLICATION_CREDENTIALS, 'r') as file:
    creds_info = json.load(file)

# 从 JSON 文件中获取 SPREADSHEET_ID
SPREADSHEET_ID = creds_info["spreadsheet_id"]

# 初始化 Google Sheets API 凭证
creds = service_account.Credentials.from_service_account_info(creds_info, scopes=SCOPES)
service = build('sheets', 'v4', credentials=creds)
sheets = service.spreadsheets()
gc = gspread.authorize(creds)
worksheet = gc.open_by_key(SPREADSHEET_ID).sheet1

# Route to serve index.html
@app.route('/')
def index():
    logging.debug("Serving index.html")
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'index.html')

# Route to handle updates from the front-end and return all sheet data
@app.route('/process_sheets', methods=['POST'])
def process_sheets():
    try:
        logging.debug("Received POST request to /process_sheets")
        data = request.get_json()
        logging.debug(f"Request data: {data}")
        update_values(data)
        result_data, dropdown_options = get_sheet_data()
        logging.debug(f"Response data: {result_data}")
        return jsonify({'data': result_data, 'dropdown_options': dropdown_options})
    except Exception as e:
        logging.error(f"Error processing request: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Route to get dropdown options
@app.route('/dropdown_options', methods=['GET'])
def dropdown_options():
    try:
        logging.debug("Received GET request to /dropdown_options")
        _, dropdown_options = get_sheet_data()
        logging.debug(f"Dropdown options: {dropdown_options}")
        return jsonify({'dropdown_options': dropdown_options})
    except Exception as e:
        logging.error(f"Error processing request: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Route to download excel
@app.route('/download_excel', methods=['GET'])
def download_excel():
    try:
        logging.debug("Received GET request to /download_excel")

        # 读取C2,C3单元格
        C2 = worksheet.acell("C2").value
        C3 = worksheet.acell("C3").value
        title = f"{C2}{C3}地铁隧道物资清单"

        # 使用定义的表头
        default_header = [
            "大类", "名称", "型号", "计量单位", "初始用量", "推进用量", "合计用量",
            "月租赁单价", "一次性单价", "安拆装单价", "月租费合价", "租赁月",
            "合价", "","","","","","备注"
        ]
        sub_header = [
            "", "", "", "", "", "", "", "", "", "", "","", "租赁系数", "租费合价",
            "一次性合价", "物流运输费", "安拆装", "合计"
        ]

        # 获取 Google Sheet 数据
        result = sheets.values().get(spreadsheetId=SPREADSHEET_ID, range="A1:S47").execute()
        all_data = result.get('values', [])

        if not all_data:
            logging.error("Excel数据为空，无法生成")
            return jsonify({'success': False, 'error': "Excel数据为空，无法生成"})

        # 提取 A16:S47 的数据
        data = all_data[15:]

        # 使用 openpyxl 生成excel
        wb = Workbook()
        sheet = wb.active

        # 添加标题行 (合并单元格，居中)
        sheet.merge_cells('A1:S1')
        title_cell = sheet['A1']
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(bold=True)

        # 添加表头
        sheet.append(default_header)
        sheet.append(sub_header)

        # 合并单元格
        for col in range(1, 13):  # A2 to L3
            sheet.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        sheet.merge_cells(start_row=2, start_column=13, end_row=2, end_column=18)  # M2 to R2
        sheet.merge_cells(start_row=2, start_column=19, end_row=3, end_column=19)  # S2 to S3

        # 设置第二行和第三行的样式
        for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=19):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

        # 添加数据
        for row_data in data:
            sheet.append(row_data)

        # 设置边框
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side('thin'))
        for row in sheet.iter_rows(min_row=1, max_row=len(data) + 3):
            for cell in row:
                cell.border = border

        # 设置数字格式，设置自动换行，调整列宽
        for row in sheet.iter_rows(min_row=4, max_row=len(data) + 3):
            for col_idx, cell in enumerate(row):
                if col_idx >= 4 and col_idx <= 17:
                    try:
                        cell.number_format = '#,##0.00'
                    except:
                        pass
                if col_idx == 18:
                    cell.alignment = Alignment(wrap_text=True)
                column_letter = get_column_letter(col_idx + 1)
                sheet.column_dimensions[column_letter].width = 15
        sheet.column_dimensions['S'].width = 40

        # 移除最后一行之后的边框
        last_row = len(data) + 3
        for col in range(1, 20):
            cell = sheet.cell(row=last_row + 1, column=col)
            cell.border = Border()

        # 设置列宽
        column_widths = {
            'A': 11.5,
            'B': 16,
            'C': 30.63,
            'D': 7.5,
            'E': 9.25,
            'F': 7.38,
            'G': 7.38,
            'H': 9.13,
            'I': 9.88,
            'J': 10.13,
            'K': 9.63,
            'L': 5.75,
            'M': 8.75,
            'N': 9.63,
            'O': 9.5,
            'P': 9.75,
            'Q': 7.88,
            'R': 9.63,
            'S': 24.88
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)

            logging.debug("Excel file generated successfully")
            return send_file(tmp.name, as_attachment=True, download_name='sheet_download.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logging.error(f"Error processing Excel download: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Fetch sheet data from Google Sheet
def get_sheet_data():
    logging.debug("Fetching data from Google Sheet")

    try:
        # Fetch metadata to get the number of rows and columns
        spreadsheet = sheets.get(spreadsheetId=SPREADSHEET_ID).execute()
        num_rows = spreadsheet.get('sheets')[0].get('properties').get('gridProperties').get('rowCount')
        num_cols = spreadsheet.get('sheets')[0].get('properties').get('gridProperties').get('columnCount')

        range = f"A1:{chr(64+num_cols)}{num_rows}"
        logging.debug(f"Fetching data from sheet with range {range}")
        result = sheets.values().get(spreadsheetId=SPREADSHEET_ID, range=range).execute()
        all_data = result.get('values', [])
        logging.debug(f"Data from sheet: {all_data}")

        # Get data validation rule for C4
        dropdown_options = []
        result = sheets.values().get(spreadsheetId=SPREADSHEET_ID, range="基础物流价格信息!A2:A10").execute()
        values = result.get("values", [])
        dropdown_options = [value[0] for value in values]
        logging.debug(f"Dropdown options: {dropdown_options}")

        return all_data, dropdown_options
    except Exception as e:
        logging.error(f"Error fetching data from Google Sheet: {e}")
        raise

# Update sheet values with user input
def update_values(data):
    logging.debug("Updating Google Sheet")
    values = []
    for key, value in data.items():
        if key == 'C4':  # 特殊处理 C4 下拉框
            value = value
        row_idx = int(key[1:]) - 1
        col_idx = ord(key[0]) - 65
        values.append({'range': key, 'values': [[value]]})
    body = {'value_input_option': 'USER_ENTERED', 'data': values}
    try:
        result = sheets.values().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        logging.debug("Update operation successful")
        return result
    except Exception as e:
        logging.error(f"Error updating Google Sheet: {e}")
        raise

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
